import subprocess

from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from .models import NextCounter
import os, base64

# Create your views here.
def index(request):
    return render(request, 'app_recorder_tts/index.html')

@csrf_exempt
def recording(request, status):




    # 스크립트 엑셀 경로 설정

    # Script Excel 파일 경로
    # os.getcwd() : 현재 경로 project_backup_tts
    # os.path.join(arg1, arg2) : arg1과 arg2의 경로 연결
    # os.listdir(transcript_list_path)[0] : transcript_list_path의 첫 번째 Excel파일
    transcript_list_path = os.path.join(os.getcwd(), 'script_list')
    transcript_list_file = os.path.join(transcript_list_path, os.listdir(transcript_list_path)[0])



    # 스크립트 엑셀 읽어오기

    # from openpyxl import load_workbook : 엑셀 읽어오기
    # data_only=Ture : 수식이 아닌 값으로 받아오기
    load_wb = load_workbook(transcript_list_file, data_only=True)

    # 시트 이름으로 불러오기
    load_ws = load_wb['Sheet1']

    # 셀 주소로 값 출력
    # print(load_ws['A1'].value)
    # 셀 좌표로 값 출력
    # print(load_ws.cell(1, 2).value)

    # 모든 행과 열 출력
    all_values = []
    for row in load_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)

    # print(all_values[0])
    # 구분 행이 삭제된 script list : [[번호, 스크립트], [번호, 스크립트], ....]
    del all_values[0]
    # print(all_values[0][1])







    # 처음, 이전, 다음 버튼 구현

    # 전체 script 수
    len_all_values = len(all_values)

    # 현재 작업 번호를 저장하는 model NextCounter{name:'cnt', numbering:0}
    model_next_counter = NextCounter.objects.filter(name='cnt')
    # print("model_next_counter : ", model_next_counter)
    # print("len(model_next_counter) : ", len(model_next_counter))

    # NextCounter 인스턴스가 없다면 생성해서 카운트 기능 구현
    if len(model_next_counter) < 1:
        new_count = NextCounter(name="cnt", numbering=1)
        new_count.save()

    # 현재 작업 번호는 NextCounter 테이블의 numbering에서 확인
    nownum = NextCounter.objects.filter(name='cnt')[0].numbering
    # print("nownum : ", nownum)



    # 초기화 버튼이 눌렸다면 model NextCounter의 numbering 0으로 변경
    if status == 'init':
        NextCounter.objects.filter(name='cnt').update(numbering = 1)
        return redirect('recording', status='now')
    # 다음 버튼이 눌렸다면 model NextCounter의 numbering 1 추가
    elif status == 'next' and nownum < len_all_values:
        NextCounter.objects.filter(name='cnt').update(numbering=(nownum + 1))
        return redirect('recording', status='now')
    # 이전 버튼이 눌렸다면 model NextCounter의 numbering 1 감소
    elif status == 'pre' and nownum > 1:
        NextCounter.objects.filter(name='cnt').update(numbering=(nownum - 1))
        return redirect('recording', status='now')

    # result_nownum 변수에 model NextCounter의 numbering 값 저장
    result_nownum = NextCounter.objects.filter(name='cnt')[0].numbering
    #print("result_nownum", result_nownum)

    show_script = all_values[result_nownum-1][1]








    wav_path = "../media/output_" + str(result_nownum) + ".wav"
    if request.method == 'POST' and status == 'record':
        save_to_server(request, result_nownum)



    if request.method == 'POST' and status == 'edit':
        cut_and_resave(request, wav_path)







    context = {'len_all_values': len_all_values, 'result_nownum': result_nownum, 'show_script': show_script, 'wav_path': wav_path}

    return render(request, 'app_recorder_tts/recording.html', context)











def save_to_server(request, result_nownum):
    # 녹음 기능 구현

    # javascript로 녹음 기능을 만들면 브라우저에서 운영체제 내의 폴더경로에 대한 접근이 불가능
    # 파일등록시에도 자바스크립트에서 파일 경로에 대한 값(파일과 폴더위치)을 가져올 수가 없어 views.py에서 저장

    #print('reuqest.POST : ', request.POST)
    base64data = request.POST['base64data']
    decode_string = base64.b64decode(base64data.split(",")[1])
    WAVE_OUTPUT_FILENAME = "./media/output_" + str(result_nownum) + ".wav"
    wav_file = open(WAVE_OUTPUT_FILENAME, "wb")
    wav_file.write(decode_string)

    return redirect('recording', status='edit')



def cut_and_resave(request, wav_path):
    # 구간 설정하면 resave wav파일
    start_sec = request.POST.get('start_sec')
    end_sec = request.POST.get('end_sec')
    wav_filepath = wav_path

    print("request.Get : ", request.POST)
    print("start_sec : ", start_sec)
    print("end_sec : ", end_sec)
    print("wav_filepath : ", wav_filepath)
    if start_sec != None and end_sec != None and wav_filepath != None:



        duration_time = float(end_sec) - float(start_sec)

        # 충돌방지 선제확인
        if 'temp_file.wav' in os.listdir('./media'):
            os.remove('./media/temp_file.wav')

        media_filepath = wav_filepath[1:]
        print("media_filepath : ", media_filepath)

        cmd = ['ffmpeg', '-y', '-i', media_filepath, '-ss', str(start_sec), '-t', str(duration_time),
               './media/temp_file.wav']
        subprocess.call(cmd)

        # ffmpeg은 같은 이름으로 덮어쓰는 과정에서 기존 파일이 손상되므로 임시 파일 활용
        os.remove(media_filepath)
        os.rename('./media/temp_file.wav', media_filepath)